from django.shortcuts import render, redirect
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import ListView, TemplateView, FormView
from django.views import View
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
import csv
import json
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from .models import ReporteExportacion, ConfiguracionReporte
from .forms import ConfiguracionReporteForm
from apps.core.models import Periodo
from apps.proyectos.models import Proyecto


class ReporteListView(LoginRequiredMixin, ListView):
    """Lista de reportes"""
    model = ReporteExportacion
    template_name = 'reportes/reporte_list.html'
    context_object_name = 'reportes'
    
    def get_queryset(self):
        return ReporteExportacion.objects.filter(usuario=self.request.user)


class ExportarView(LoginRequiredMixin, View):
    """Vista de exportación"""
    template_name = 'reportes/exportar.html'
    
    def get(self, request):
        from .forms import ExportarReporteForm
        
        # Cargar configuración del usuario
        config = None
        initial_data = {}
        
        try:
            config = ConfiguracionReporte.objects.get(usuario=request.user)
            initial_data = {
                'formato': config.formato_exportacion,
                'separador': config.separador_csv,
                'incluir_totales': True,
                'incluir_descripcion': True,
            }
        except ConfiguracionReporte.DoesNotExist:
            pass
            
        form = ExportarReporteForm(initial=initial_data, user=request.user)
        context = {
            'form': form,
            'periodos': Periodo.objects.filter(usuario=request.user),
            'proyectos': Proyecto.objects.filter(usuario=request.user),
            'config': config
        }
        return render(request, self.template_name, context)
    
    def post(self, request):
        from .forms import ExportarReporteForm
        from apps.horas.models import RegistroHora
        
        form = ExportarReporteForm(request.POST, user=request.user)
        
        # Si es preview
        if request.POST.get('preview'):
            if form.is_valid():
                # Obtener datos filtrados
                registros = self.get_filtered_data(request.user, form.cleaned_data)
                preview_html = self.generate_preview_html(registros[:10])  # Solo primeros 10
                return JsonResponse({
                    'success': True,
                    'preview_html': preview_html,
                    'total_records': registros.count()
                })
            else:
                return JsonResponse({
                    'success': False,
                    'message': 'Error en los filtros'
                })
        
        # Si es exportación
        if form.is_valid():
            registros = self.get_filtered_data(request.user, form.cleaned_data)
            formato = form.cleaned_data.get('formato', 'csv')
            
            if formato == 'csv':
                return self.export_csv(registros, form.cleaned_data)
            elif formato == 'xlsx':
                return self.export_xlsx(registros, form.cleaned_data)
            
        return JsonResponse({'success': False, 'message': 'Error en el formulario'})
    
    def get_filtered_data(self, user, data):
        from apps.horas.models import RegistroHora
        
        queryset = RegistroHora.objects.filter(usuario=user)
        
        if data.get('fecha_inicio'):
            queryset = queryset.filter(fecha__gte=data['fecha_inicio'])
        if data.get('fecha_fin'):
            queryset = queryset.filter(fecha__lte=data['fecha_fin'])
        if data.get('periodo'):
            queryset = queryset.filter(fecha__range=[data['periodo'].fecha_inicio, data['periodo'].fecha_fin])
        if data.get('proyectos'):
            queryset = queryset.filter(proyecto__in=data['proyectos'])
            
        return queryset.select_related('proyecto').order_by('fecha', 'proyecto__nombre')
    
    def generate_preview_html(self, registros):
        if not registros:
            return '<div class="alert alert-info">No se encontraron registros</div>'
        
        html = '<div class="table-responsive"><table class="table table-sm">'
        html += '<thead><tr><th>Fecha</th><th>Proyecto</th><th>Horas</th><th>Tipo</th></tr></thead><tbody>'
        
        for registro in registros:
            html += f'<tr><td>{registro.fecha}</td><td>{registro.proyecto.nombre}</td><td>{registro.horas}</td><td>{registro.tipo_tarea}</td></tr>'
        
        html += '</tbody></table></div>'
        return html
    
    def export_csv(self, registros, data):
        from datetime import date
        
        response = HttpResponse(content_type='text/csv')
        fecha_actual = date.today().strftime('%Y-%m-%d')
        response['Content-Disposition'] = f'attachment; filename="reporte_horas_{fecha_actual}.csv"'
        
        separador = data.get('separador', ',')
        writer = csv.writer(response, delimiter=separador)
        
        # Obtener configuración del usuario
        try:
            config = ConfiguracionReporte.objects.get(usuario=self.request.user)
            separador_decimal = getattr(config, 'separador_decimal', '.')
        except ConfiguracionReporte.DoesNotExist:
            separador_decimal = '.'
        
        # Encabezados
        writer.writerow(['Fecha', 'Proyecto', 'Cliente', 'Horas', 'Tipo Tarea', 'Descripcion'])
        
        # Datos
        for registro in registros:
            horas_str = str(registro.horas).replace('.', separador_decimal)
            writer.writerow([
                registro.fecha,
                registro.proyecto.nombre,
                registro.proyecto.cliente or '',
                horas_str,
                registro.tipo_tarea,
                registro.descripcion or ''
            ])
        
        return response
    
    def export_xlsx(self, registros, data):
        from datetime import date
        
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return JsonResponse({'success': False, 'message': 'openpyxl no está instalado'})
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Reporte de Horas"
        
        # Encabezados
        headers = ['Fecha', 'Proyecto', 'Cliente', 'Horas', 'Tipo Tarea', 'Descripcion']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Datos
        for row, registro in enumerate(registros, 2):
            ws.cell(row=row, column=1, value=registro.fecha)
            ws.cell(row=row, column=2, value=registro.proyecto.nombre)
            ws.cell(row=row, column=3, value=registro.proyecto.cliente or '')
            ws.cell(row=row, column=4, value=float(registro.horas))
            ws.cell(row=row, column=5, value=registro.tipo_tarea)
            ws.cell(row=row, column=6, value=registro.descripcion or '')
        
        fecha_actual = date.today().strftime('%Y-%m-%d')
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="reporte_horas_{fecha_actual}.xlsx"'
        wb.save(response)
        
        return response


class ConfiguracionView(LoginRequiredMixin, FormView):
    """Vista de configuración de reportes"""
    template_name = 'reportes/configuracion.html'
    form_class = ConfiguracionReporteForm
    success_url = '/reportes/configuracion/'
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['usuario'] = self.request.user
        return kwargs
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['periodos'] = Periodo.objects.filter(usuario=self.request.user)
        context['proyectos'] = Proyecto.objects.filter(usuario=self.request.user)
        return context
    
    def form_valid(self, form):
        form.save()
        messages.success(self.request, 'Configuración guardada correctamente.')
        return super().form_valid(form)


# API Views
class ExportarCSVAPIView(APIView):
    """API exportar CSV"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        # Aquí iría la lógica de exportación CSV
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="horas.csv"'
        response.write('fecha,proyecto,horas,tipo_tarea\n')
        return response


class ExportarXLSXAPIView(APIView):
    """API exportar XLSX"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        return Response({'message': 'Exportar XLSX'})


class HistorialExportacionAPIView(APIView):
    """API historial de exportaciones"""
    permission_classes = [IsAuthenticated]
    
    def get(self, request):
        exportaciones = ReporteExportacion.objects.filter(usuario=request.user)
        data = [{
            'id': e.id,
            'nombre_archivo': e.nombre_archivo,
            'formato': e.formato,
            'created_at': e.created_at
        } for e in exportaciones]
        return Response(data)
